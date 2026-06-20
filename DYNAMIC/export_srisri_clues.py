from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(__file__).with_name("SriSri_Puzzle_Counts_TextRows.xlsm")
OUTPUT_DIR = Path(__file__).with_name("SS")


def normalize_value(value):
    if isinstance(value, str):
        value = value.strip()
        return value or None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def row_to_record(row):
    values = [normalize_value(cell) for cell in row[:4]]
    while len(values) < 4:
        values.append(None)
    return {
        "crossNumber": values[0],
        "crossText": values[1],
        "downNumber": values[2],
        "downText": values[3],
    }


def has_content(record):
    return any(value is not None for value in record.values())


def build_sheet_payload(sheet_name, worksheet):
    cross_entries = []
    down_entries = []

    for row_index, row in enumerate(
        worksheet.iter_rows(values_only=True),
        start=1,
    ):
        record = row_to_record(row)
        if not has_content(record):
            continue

        if record["crossNumber"] is not None or record["crossText"] is not None:
            cross_entries.append(
                {
                    "row": row_index,
                    "number": record["crossNumber"],
                    "text": record["crossText"],
                }
            )

        if record["downNumber"] is not None or record["downText"] is not None:
            down_entries.append(
                {
                    "row": row_index,
                    "number": record["downNumber"],
                    "text": record["downText"],
                }
            )

    return {
        "sheetName": sheet_name,
        "identity": f"{sheet_name}C",
        "crossEntries": cross_entries,
        "downEntries": down_entries,
    }


def main():
    OUTPUT_DIR.mkdir(exist_ok=True)

    workbook = load_workbook(
        WORKBOOK_PATH,
        data_only=True,
        read_only=True,
        keep_vba=True,
    )

    created_files = []

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        payload = build_sheet_payload(sheet_name, worksheet)
        output_path = OUTPUT_DIR / f"{sheet_name}C.json"
        output_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8-sig",
        )
        created_files.append(output_path.name)

    print(f"Created {len(created_files)} clue JSON files in {OUTPUT_DIR}")
    for name in created_files:
        print(name)


if __name__ == "__main__":
    main()
